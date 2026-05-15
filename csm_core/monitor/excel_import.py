"""Excel batch-import parser for monitor tasks.

Schema (fixed columns, Chinese headers — what the user actually pastes
into a spreadsheet):

  | 类型      | 名称        | URL                    | 关键词        | TopN | 调度  |
  |-----------|-------------|------------------------|---------------|------|-------|
  | 知乎问题  | 任务A       | https://www.zhihu.com…  | 我的品牌      | 10   | 09:00 |
  | B站评论   | 任务B       | https://www.bilibili.com…| 自发评论原文 | 20   | manual |

- ``类型`` accepts both the human label (``知乎问题``, ``B站评论``,
  ``抖音评论``, ``快手评论``) and the underlying TaskType code
  (``zhihu_question`` etc.) — whichever is more convenient.
- ``关键词`` is the brand keyword for Zhihu, the self-published comment
  text for the comment platforms. Same column slot, type-dependent
  semantics.
- ``TopN`` defaults to 10 if blank.
- ``调度`` accepts ``manual`` or ``HH:MM``; blank == ``manual``.

The parser is intentionally lenient: it returns ``ImportReport`` with
both successfully-parsed tasks and per-row error messages so the UI can
let the user fix bad rows without rejecting the whole file.

Why a separate parser instead of inlining in the page: the spreadsheet
format is exactly what we want to unit-test (column header detection,
type-label mapping, validation), and exposing it via plain-Python
functions means tests don't need a QApplication.
"""
from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from .base import MonitorTask, TaskType


# Mapping from human-readable label → TaskType code. The TaskType code
# itself is also accepted (we lowercase + check the dict and fall back
# to direct match) so power users can paste codes directly.
_TYPE_LABEL_MAP: dict[str, TaskType] = {
    "知乎问题": "zhihu_question",
    "知乎": "zhihu_question",
    "zhihu": "zhihu_question",
    "zhihu_question": "zhihu_question",
    "b站评论": "bilibili_comment",
    "b 站评论": "bilibili_comment",
    "b站": "bilibili_comment",
    "bilibili": "bilibili_comment",
    "bilibili_comment": "bilibili_comment",
    "抖音评论": "douyin_comment",
    "抖音": "douyin_comment",
    "douyin": "douyin_comment",
    "douyin_comment": "douyin_comment",
    "快手评论": "kuaishou_comment",
    "快手": "kuaishou_comment",
    "kuaishou": "kuaishou_comment",
    "kuaishou_comment": "kuaishou_comment",
    "百度关键词": "baidu_keyword",
    "百度": "baidu_keyword",
    "baidu": "baidu_keyword",
    "baidu_keyword": "baidu_keyword",
}

# Canonical column headers we look for. The parser tolerates extra
# whitespace and different cases; ordering inside the file is
# irrelevant — we resolve columns by header name.
_HEADERS = {
    "type": ["类型", "type"],
    "name": ["名称", "name"],
    "url": ["url", "链接", "目标url", "目标链接"],
    "keyword": ["关键词", "品牌词", "自发评论", "评论文本", "keyword"],
    "top_n": ["topn", "top_n", "top n", "top-n", "范围"],
    "schedule": ["调度", "schedule", "时间", "定时"],
}


# Column headers exposed to the user when they download the template.
TEMPLATE_HEADERS = ["类型", "名称", "URL", "关键词", "TopN", "调度"]
TEMPLATE_SAMPLES = [
    ["知乎问题", "知乎-某品牌词监测", "https://www.zhihu.com/question/123456", "ACME", 10, "09:00"],
    ["B站评论", "B站-某视频评论留存", "https://www.bilibili.com/video/BV1xxxxx", "你这个测评太真实了", 20, "manual"],
    ["抖音评论", "抖音-某视频评论留存", "https://www.douyin.com/video/7300000000000000000", "支持博主", 10, "manual"],
    ["快手评论", "快手-某视频评论留存", "https://www.kuaishou.com/short-video/3xxxxxxxx", "已加购物车", 10, "manual"],
    ["百度关键词", "百度-Claude教程", "search:Claude Code 教程", "Claude|Anthropic", 10, "09:00"],
]


@dataclass
class ImportReport:
    """Outcome of parsing one Excel file.

    ``tasks`` and ``errors`` are independent — a file may yield both
    valid tasks and per-row errors, and the UI typically saves the
    former while showing the latter so the user can fix and retry.
    """

    tasks: list[MonitorTask] = field(default_factory=list)
    # Each entry is (row_number, message). Row numbers are 1-based and
    # match what Excel shows (header is row 1, first data row is 2).
    errors: list[tuple[int, str]] = field(default_factory=list)

    @property
    def ok_count(self) -> int:
        return len(self.tasks)

    @property
    def error_count(self) -> int:
        return len(self.errors)


# ── Public API ──────────────────────────────────────────────────────────────
def parse_excel(path: str | Path) -> ImportReport:
    """Open ``path`` (xlsx) and parse its first sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is required for Excel batch import") from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: list[list] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
    finally:
        wb.close()

    return parse_rows(rows)


def parse_rows(rows: list[list]) -> ImportReport:
    """Parse pre-extracted spreadsheet rows.

    Splitting this from ``parse_excel`` lets unit tests exercise the
    schema logic without writing real .xlsx files to tmp dirs.
    """
    report = ImportReport()
    if not rows:
        return report

    header_idx = _resolve_headers(rows[0])
    missing = [k for k in ("type", "url", "keyword") if k not in header_idx]
    if missing:
        report.errors.append((1, f"缺少必需列：{', '.join(missing)}（参考模板）"))
        return report

    for excel_row, raw in enumerate(rows[1:], start=2):
        if not raw or all(_is_blank(c) for c in raw):
            continue  # skip blank rows
        try:
            task = _row_to_task(raw, header_idx)
            report.tasks.append(task)
        except ValueError as e:
            report.errors.append((excel_row, str(e)))
    return report


def write_template(path: str | Path) -> None:
    """Drop a fresh template .xlsx at ``path`` for the user to fill in."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "监测任务批量导入"

    # Header row — bold + light fill so it visually separates from data.
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EFEAE0")
    for col, label in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, sample in enumerate(TEMPLATE_SAMPLES, start=2):
        for col, value in enumerate(sample, start=1):
            ws.cell(row=row_idx, column=col, value=value)

    # Reasonable column widths — eyeballed for typical content.
    widths = [12, 28, 56, 30, 8, 10]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # NOTE: we deliberately do NOT add a "说明" row inside the data
    # range — the parser would have to special-case skipping it and
    # users editing the template might accidentally treat it as part
    # of the data table. Format guidance lives in the import-dialog
    # body text instead, where it's reachable without opening Excel.

    wb.save(str(path))


# ── Internal ────────────────────────────────────────────────────────────────
def _resolve_headers(header_row: list) -> dict[str, int]:
    """Map our internal column key → 0-based index in the spreadsheet."""
    idx: dict[str, int] = {}
    for col, raw in enumerate(header_row):
        norm = _normalize_header(raw)
        if not norm:
            continue
        for key, candidates in _HEADERS.items():
            if norm in candidates:
                idx.setdefault(key, col)
                break
    return idx


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("　", "").replace(" ", "")


def _row_to_task(row: list, header_idx: dict[str, int]) -> MonitorTask:
    type_raw = _cell(row, header_idx.get("type"))
    name_raw = _cell(row, header_idx.get("name"))
    url_raw = _cell(row, header_idx.get("url"))
    keyword_raw = _cell(row, header_idx.get("keyword"))
    top_n_raw = _cell(row, header_idx.get("top_n"))
    schedule_raw = _cell(row, header_idx.get("schedule"))

    type_norm = _normalize_header(type_raw)
    if not type_norm:
        raise ValueError("类型列为空")
    ttype = _TYPE_LABEL_MAP.get(type_norm)
    if not ttype:
        raise ValueError(f"未识别的类型「{type_raw}」")

    if not url_raw:
        raise ValueError("URL 为空")
    url_text = str(url_raw).strip()
    # Allow both http(s) URLs and baidu's "search:" prefix
    if not (url_text.startswith("http") or url_text.startswith("search:")):
        raise ValueError("URL 需以 http(s) 或 search: 开头")
    if not keyword_raw:
        raise ValueError("关键词/自发评论为空")

    # Default name from the URL if the user didn't bother to fill it.
    name = (str(name_raw).strip() if name_raw else "") or _default_name(ttype, str(url_raw))

    top_n = _parse_int(top_n_raw, default=10)
    if top_n < 1:
        raise ValueError("TopN 需为正整数")

    schedule = _parse_schedule_cell(schedule_raw)

    config: dict[str, object] = {"top_n": top_n}
    if ttype == "zhihu_question":
        config["target_brand"] = str(keyword_raw).strip()
    elif ttype == "baidu_keyword":
        # 百度：「关键词」列里放「BrandA|BrandB|...」，URL 列填 "search:实际搜索词"。
        # adapter 内部 fetch 时从 config.search_keyword 拼真实 URL，
        # 表里的 target_url 只是占位。
        url_text = str(url_raw).strip()
        if url_text.startswith("search:"):
            search_keyword = url_text[len("search:"):].strip()
        else:
            search_keyword = url_text  # 容错：直接当关键词
        if not search_keyword:
            raise ValueError("百度任务的搜索关键词为空")
        brands_raw = str(keyword_raw).strip()
        brands = [b.strip() for b in brands_raw.split("|") if b.strip()]
        if not brands:
            raise ValueError("百度任务的目标品牌词为空")
        from urllib.parse import quote as _quote
        return MonitorTask(
            type=ttype,
            name=name,
            target_url=f"https://www.baidu.com/s?wd={_quote(search_keyword)}",
            config={"search_keyword": search_keyword, "target_brands": brands},
            schedule_cron=schedule,
            enabled=True,
        )
    else:
        config["my_comment_text"] = str(keyword_raw).strip()

    return MonitorTask(
        type=ttype,
        name=name,
        target_url=str(url_raw).strip(),
        config=config,
        schedule_cron=schedule,
        enabled=True,
    )


def _cell(row: list, idx: int | None):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _is_blank(cell) -> bool:
    if cell is None:
        return True
    if isinstance(cell, str) and not cell.strip():
        return True
    return False


def _parse_int(value, *, default: int) -> int:
    if value is None or (isinstance(value, str) and not value.strip()):
        return default
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).strip())
    except ValueError:
        raise ValueError(f"TopN 不是数字：{value}")


def _parse_schedule_cell(value) -> str:
    """Accept blank → 'manual', 'manual', or HH:MM. Reject anything else."""
    if value is None:
        return "manual"
    text = str(value).strip()
    if not text or text.lower() == "manual":
        return "manual"
    # Excel sometimes hands us a datetime.time when the cell is
    # formatted as time. Render it back to HH:MM.
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return f"{value.hour:02d}:{value.minute:02d}"
    if ":" in text:
        try:
            hh, mm = text.split(":", 1)
            h, m = int(hh), int(mm)
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        except ValueError:
            pass
    raise ValueError(f"调度格式错误（应为 manual 或 HH:MM）：{text}")


def _default_name(ttype: TaskType, url: str) -> str:
    label = {
        "zhihu_question": "知乎问题",
        "bilibili_comment": "B站评论",
        "douyin_comment": "抖音评论",
        "kuaishou_comment": "快手评论",
    }.get(ttype, ttype)
    # Trim the URL to a recognizable suffix (last 24 chars).
    tail = url[-24:] if len(url) > 24 else url
    return f"{label} · {tail}"
